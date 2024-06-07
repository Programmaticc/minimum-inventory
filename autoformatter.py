from openpyxl import load_workbook
import pandas as pd
import os
import glob

def copy_first_line_to_top(input_file):
    wb = load_workbook(input_file)
    ws = wb.active
    first_row_data = [cell.value for cell in ws[1]]
    ws.insert_rows(idx=1, amount=1)
    for col, value in enumerate(first_row_data, start=1):
        ws.cell(row=1, column=col, value=value)
    wb.save(input_file)

def find_article_column(df):
    for col in df.columns:
        if isinstance(col, str) and col.startswith('M') and len(col) > 1 and col[1].isdigit():
            return col
    return None

def find_depot_column(df):
    for col in df.columns:
        if isinstance(col, str) and col.startswith(('01', '02', '03', '04')):
            return col
    return None

def find_date_column(df):
    for col in df.columns:
        if all(isinstance(val, int) and len(str(val)) == 4 and str(val).startswith('20') for val in df[col].dropna()):
            unique_years = df[col].unique()
            if len(unique_years) > 1:
                return col
    return None

def find_numeric_columns(df, start_column):
    columns = df.columns.tolist()
    start_index = columns.index(start_column) + 1
    return columns[start_index:start_index+12]

def format_excel_files(directory):
    excel_files = glob.glob(os.path.join(directory, '*.xlsx'))
    if not excel_files:
        print(f"No Excel files found in directory: {directory}")
        return

    for input_file in excel_files:
        output_file = os.path.join(directory, os.path.splitext(os.path.basename(input_file))[0] + '_output.xlsx')

        try:
            copy_first_line_to_top(input_file)

            df = pd.read_excel(input_file)
            first_empty_row = df.isnull().all(axis=1).idxmax()
            df = pd.read_excel(input_file, skiprows=range(1, first_empty_row + 1))

            article_col = find_article_column(df)
            depot_col = find_depot_column(df)
            date_col = find_date_column(df)

            if article_col and depot_col and date_col:
                sales_cols = find_numeric_columns(df, date_col)
                keep_columns = [article_col, depot_col, date_col] + sales_cols
                df_filtered = df[keep_columns]

                df_filtered.to_excel(output_file, index=False, header=False)
                print(f"Excel file formatted and saved successfully to: {output_file}")
            else:
                print(f"Error: Could not find all required columns in file '{input_file}' based on the specified criteria.")

        except FileNotFoundError:
            print(f"Error: Input file '{input_file}' not found.")
        except pd.errors.ParserError:
            print(f"Error: Could not parse Excel file '{input_file}'.")
        except Exception as e:
            print(f"An unexpected error occurred while processing file '{input_file}': {e}")

# Example usage
directory = r"C:\Users\Desktop"
format_excel_files(directory)


